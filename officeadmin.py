from openpyxl import *
from tkinter import *
from tkinter import messagebox
  

wbook = load_workbook('C:\\Users\\Admin\\AppData\\Local\\Programs\\Python\\Python39\\Nandha my_assistant\\schedule.xlsx') 
      
sheet = wbook.active 
      
      
def excel(): 
          
       
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
            
    sheet.cell(row=1, column=1).value = "Event Name"
    sheet.cell(row=1, column=2).value = "Date"
    sheet.cell(row=1, column=3).value = "Place"
    sheet.cell(row=1, column=4).value = "Time"
       

def focus1(event): 
       
    EventName_field.focus_set() 
      
      
    # Function to set focus 
def focus2(event): 
       
    Date_field.focus_set() 
      
      
     
def focus3(event): 
        
        Place_field.focus_set() 
      
      
     
def focus4(event): 
       
        Time_field.focus_set() 
      
      

      

def clear(): 
          
     
        EventName_field.delete(0, END) 
        Date_field.delete(0, END) 
        Place_field.delete(0, END) 
        Time_field.delete(0, END) 
      
      
      

def insert(): 
          
         
        if(dt.get()=="159"):
            current_row = sheet.max_row 
            current_column = sheet.max_column 
      
           
            sheet.cell(row=current_row + 1, column=1).value = EventName_field.get() 
            sheet.cell(row=current_row + 1, column=2).value = Date_field.get() 
            sheet.cell(row=current_row + 1, column=3).value = Place_field.get() 
            sheet.cell(row=current_row + 1, column=4).value = Time_field.get() 
           
            wbook.save('C:\\Users\\Admin\\AppData\\Local\\Programs\\Python\\Python39\\Nandha my_assistant\\schedule.xlsx') 
      
            # set focus on the name_field box 
            EventName_field.focus_set() 
      
            # call the clear() function 
            clear()
            root.destroy()
        else:
            messagebox.showinfo("warning","Please enter a correct password")
      
      


          
        # create a GUI window
        
root = Tk() 
          
# set the background colour of GUI window 
root.configure(background='light blue') 
# Adjust size 
root.geometry("500x300")
root.title("Events") 
          
# set the configuration of GUI window 
root.geometry("500x300") 
          
excel() 
          
# create a Form label 
heading = Label(root, text="EVENTS", bg="white") 
#create password label
pwd=Label(root,text="Password",bg="light blue")
            
# create a Name label 
EventName = Label(root, text="EventName", bg="light blue") 
          
# create a Course label 
Date = Label(root, text="Date", bg="light blue") 
          
# create a Semester label 
Place = Label(root, text="Place", bg="light blue") 
          
# create a Form No. lable 
Time = Label(root, text="Time", bg="light blue") 
heading.grid(row=0, column=1)
pwd.grid(row=1,column=0)
EventName.grid(row=2, column=0) 
Date.grid(row=3, column=0) 
Place.grid(row=4, column=0) 
Time.grid(row=5, column=0) 
dt=StringVar()
pwd_field=Entry(root,textvariable=dt)
EventName_field = Entry(root) 
Date_field = Entry(root) 
Place_field = Entry(root) 
Time_field = Entry(root)
EventName_field.bind("<Return>", focus1)
Date_field.bind("<Return>", focus2) 
           
Place_field.bind("<Return>", focus3) 
          
            
Time_field.bind("<Return>", focus4) 
          
pwd_field.grid(row=1,column=1,ipadx="100")
EventName_field.grid(row=2, column=1, ipadx="100") 
Date_field.grid(row=3, column=1, ipadx="100") 
Place_field.grid(row=4, column=1, ipadx="100") 
Time_field.grid(row=5, column=1, ipadx="100") 
            
excel() 
          
             
submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert) 
submit.grid(row=8, column=1) 
          
            # start the GUI

def initi():        
    root.mainloop() 
